# -*- coding: utf-8 -*-

"""
Скрипт для автоматического сбора данный с тригеров сетевого оборудование и сервисов
Рабочий файл: отчет_NetworkAvailability.xlsx - Должен быть создан по шаблону

Переменные, которые настраиваются для корректной работы скрипта:
path - путь до рабочего файла
path_logfile - путь до файла с логами
path_copy - путь, куда копируется заполненый файл
zabbix_server_url - адрес мониторинга Zabbix по WEB API
zabbix_login - Логин пользователя Zabbix по WEB API
zabbix_password - Пароль пользователя Zabbix по WEB API

@ Автор скрипта:
Зенкин Денис

Дата создания скрипта:
20.09.2021
"""

from datetime import datetime
import time

from openpyxl import load_workbook
from pyzabbix import ZabbixAPI

# Настраиваемые переменные
path = "E:\\Techical\\Network\\отчет_NetworkAvailability.xlsx"
path_logfile = "E:\\Techical\\Network\\log.txt"
zabbix_server_url = "http://zabbix40.ak.local"
zabbix_login = "zdm"
zabbix_password = "1q3hpEtv144"
path_copy = ""  # Копирование пока отключено


def log(text):
    """
        Записывает ошибку в лог файл.
    """
    file = open(path_logfile, "a")
    file.write(text)
    file.close()


def data():
    """
        Определение координат последней записи.
        Цикл проверяет столбец A, находит номер следующей пустой строки и записывает туда текущую дату.
    """
    global coord
    for i in range(1, 1000):
        y = sheet_ranges.cell(row=i, column=1).value
        if i > 5 and (y is None):
            coord = sheet_ranges.cell(row=i, column=1).row
            coord = str(coord)
            coord_last = "A" + coord
            sheet_ranges[coord_last] = date  # Определили номер строки
            break
    return coord


def move(letter):
    """
        Получаем следующую букву по алфавиту
    """
    if letter.lower() == 'z':
        return 'A' + chr(97).upper()
    if len(letter) == 2:
        a = list(letter)
        if a[1] == 'z':
            return chr(98).upper() + chr(97).upper()
        return a[0].upper() + chr(ord(a[1]) + 1).upper()
    return chr(ord(letter.lower()) + 1).upper()


def network(group):
    """
        Получает и записывает значения name и availability в таблицу
        Записывает в excel.
    """
    # Текущеее время в UnixStamp
    realtime = datetime.now()
    unixtime_now = time.mktime(realtime.timetuple())
    ut_24 = int(unixtime_now) - 86400  # Минус 24 часа от текущего времени

    # Получаем список проблем по группе
    list_event = z.event.get(
        groupids=group,
        time_from=ut_24,
        selectHosts=['host'],
    )

    # Подставляем триггер из группы
    if group['ID'] == 28:
        objectid = 23488
        list_event = z.event.get(
            groupids=group,
            time_from=ut_24,
            objectids=objectid,
            selectHosts=['host'],
        )

    devices_list = {}

    for item in list_event:
        event_next_id = item['r_eventid']
        if int(event_next_id) == 0:
            continue
        for next_item in list_event:
            if event_next_id == next_item['eventid']:
                start_event_clock = next_item['clock']
                break
        event_duration = int(start_event_clock) - int(item['clock'])
        ratio = round((event_duration / 86400 * 100), 2)
        if item['hosts'][0]['host'] in devices_list:
            devices_list[item['hosts'][0]['host']] += round(ratio, 2)
            devices_list[item['hosts'][0]['host'] + '_count'] += 1
        else:
            devices_list[item['hosts'][0]['host']] = round(ratio, 2)
            devices_list[item['hosts'][0]['host']+'_count'] = 1
        # print(item['hosts'][0]['host'], ratio)

    hosts = z.host.get(groupids=group['ID'], output=['hostid', 'name', 'host'])  # Обработка группы
    for host in hosts:
        name = host['name']  # Имя узла
        ip = host['host']
        for cellObj in sheet_ranges['A2':'AX2']:
            for cell in cellObj:
                if cell.value == ip:
                    coord_notAVL = cell.column_letter + coord
                    coord_AVL = move(cell.column_letter) + coord
                    coord_name = cell.column_letter + "3"  # Координаты строки имени
                    sheet_ranges[coord_name] = name
                    if ip not in devices_list:
                        sheet_ranges[coord_notAVL] = 0.0
                        sheet_ranges[coord_AVL] = 100
                        break
                    else:
                        sheet_ranges[coord_notAVL] = devices_list[ip+'_count']
                        sheet_ranges[coord_AVL] = 100 - devices_list[ip]


def row_1(row):
    """
        На 1 ячейку вверх.
    """
    row2 = int(row) - 1
    row2 = str(row2)
    return row2


date = datetime.today().strftime("%d.%m.%y")  # Текущая дата

file = open(path_logfile, "a")
file.write("-----------------------------------------------------------")
file.write("\n\n")
file.write("Начало работы в ")
file.write(str(datetime.now()))
file.write("\n\n")
file.close()

serial_zabbix = []

try:
    wb = load_workbook(path)  # Загружаем файл
except FileNotFoundError:
    log("Не возможно найти файл excel\n\n")
    raise
try:
    z = ZabbixAPI(zabbix_server_url)
    z.login(zabbix_login, zabbix_password)
    log("Connected to Zabbix API Version %s" % z.api_version())
    log("\n\n")
except OSError:
    log("Нет соединения с сервером Zabbix\n\n")
    raise

dist_book = {
    'Network': {'ID': 24, 'cord_summa': None},
    'Services': {'ID': 28, 'cord_summa': None}
}

for book in dist_book:
    sheet_ranges = wb[book]
    summa_prom = []
    summa = []
    coord = data()
    coord3 = "A" + coord
    coord4 = "AX" + coord
    network(dist_book[book])

try:
    wb.save(path)
    log("Запись завершена\n\n")
except PermissionError:
    log("Не возможно сохранить файл excel, у кого-то он открыт! Сохранение в промежуточный файл!\n\n")
    raise

# временно отключено, до запуска в автоматизацию
# try:
#     shutil.copy(path, path_copy)  # !!!!!!!!
# except PermissionError:
#     log("ПЗДЦ!!!! Файл не скопировался, видимо потому что он открыт у кого-то! Копирование в общую папку!\n\n")
#     raise
