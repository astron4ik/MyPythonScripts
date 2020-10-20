# -*- coding: utf-8 -*-

"""
Скрипт для автоматического сбора данный с принтеров
Рабочий файл: Отчет_принтеры.xlsx - Должен быть создан по шаблону

Переменные, которые настраиваются для корректной работы скрипта:
path - путь до рабочего файла
path_logfile - путь до файла с логами
path_copy - путь, куда копируется заполненый файл
zabbix_server_url - адрес мониторинга Zabbix по WEB API
zabbix_login - Логин пользователя Zabbix по WEB API
zabbix_password - Пароль пользователя Zabbix по WEB API

@ Автор скрипта:
Зенкин Денис

Дата создания скрипта:
06.09.2020
"""

import datetime
import shutil

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pyzabbix import ZabbixAPI

# Настраиваемые переменные
path = "<path>\\Ежедневный_отчет_по_принтерам.xlsx"
path_logfile = "<path>\\logfile.txt"
zabbix_server_url = "http://<url>"
zabbix_login = "<login>"
zabbix_password = "<password>"
path_copy = ""  # Копирование пока отключено


def log(text):
    """
        Записывает ошибку в лог файл.
    """
    file = open(path_logfile, "a")
    file.write(text)
    file.close()


def data():
    """
        Определение координат последней записи.
        Цикл проверяет столбец A, находит номер следующей пустой строки и записывает туда текущую дату.
    """
    global coord
    for i in range(1, 1000):
        y = sheet_ranges.cell(row=i, column=1).value
        if i > 7 and (y is None):
            coord = sheet_ranges.cell(row=i, column=1).row
            coord = str(coord)
            coord_last = "A" + coord
            sheet_ranges[coord_last] = date  # Определили номер строки
            break
    return coord


def total_pages_and_serial(d, number):
    """
        Определяет серийный номер и число напечатанных страниц.
        Получает и обрабатывает данные из Zabbix.
    """
    items = z.item.get(hostids=d, output=['itemid', 'name'])  # Это id принтера
    d = items[number]
    d = d.get('itemid')
    a = z.item.get(itemids=d, output=['lastvalue'])  # d это id total pages
    b = a[0]
    b = b.get('lastvalue')
    return b


def printers(group):
    """
        Получает и записывает значения serial и total pages для каждого принтера.
        Записывает в excel.
    """
    hosts = z.host.get(groupids=group['ID'], output=['hostid', 'name', 'host'])  # Группа Принтеры Офис
    for host in hosts:
        name = host['name']  # Имя узла принтера
        ip = host['host']
        # print(host['hostid'],host['name'])
        host = host.get('hostid')
        pages = int(total_pages_and_serial(host, 6))  # Узнаем Кол-во Страниц, 6 это номер позиции total pages
        serial = total_pages_and_serial(host, 2)  # Узнаем Серийник, 2 это номер позиции serial
        serial_zabbix.append(serial)
        for cellObj in sheet_ranges['A2':'DR2']:
            for cell in cellObj:
                if cell.value == serial:
                    column = cell.column_letter
                    coord_pages = cell.column_letter + coord
                    coord_name = cell.column_letter + "4"  # Координаты строки места принтера
                    coord_ip = cell.column_letter + "3"  # Координаты строки ip адреса
                    sheet_ranges[coord_name] = name
                    sheet_ranges[coord_pages] = pages
                    sheet_ranges[coord_ip] = str(ip)  # Записываем IP
                    # print(serial, ' - ', pages)
                    dict_excel[cell.value] = column
                elif cell.value != serial:  # Создает словарь из серийников, для сравнения
                    # column1 = cell.column_letter
                    dict_excel[cell.value] = cell.coordinate[:-1:]


def all_value_result(string):
    """
        Функция для записи в ячейки значений из словаря
        *** Сооброзить на свежую голову ***
    """
    for cellObj in sheet_ranges['A1':'K4']:
        for cell in cellObj:
            if string.startswith('=' + cell.value):
                coord_value = cell.column_letter + coord
                sheet_ranges[coord_value] = string
                summa_prom.append(coord_value)
                return


def raschet(cell, row_2):
    """
        Расчет для столбца прирост.
        Создает переменную с формулой расчета прироста total_pages.
    """
    # column_1 = column_index_from_string(coord_column1)
    column_1 = int(coord_column1) - 1
    column_2 = get_column_letter(column_1)
    coordinate_last = column_2 + row_2
    coordinate_now = column_2 + str(row)
    letter = "=" + coordinate_now + "-" + coordinate_last
    sheet_ranges[coordinate] = letter
    summa_prom.append(cell.coordinate)
    return summa


def row_1(row):
    """
        На 1 ячейку вверх.
    """
    row2 = int(row) - 1
    row2 = str(row2)
    return row2


date = datetime.datetime.today().strftime("%d.%m.%y")  # Текущая дата

file = open(path_logfile, "a")
file.write("-----------------------------------------------------------")
file.write("\n\n")
file.write("Начало работы в ")
file.write(str(datetime.datetime.now()))
file.write("\n\n")
file.close()

serial_zabbix = []

try:
    wb = load_workbook(path)  # Загружаем файл
except FileNotFoundError:
    log("Не возможно найти файл excel\n\n")
    raise
try:
    z = ZabbixAPI(zabbix_server_url)
    z.login(zabbix_login, zabbix_password)
    log("Connected to Zabbix API Version %s" % z.api_version())
    log("\n\n")
except OSError:
    log("Нет соединения с сервером Zabbix\n\n")
    raise

dist_book = {
    'Офис': {'ID': 45, 'cord_summa': None},
    'Склад': {'ID': 44, 'cord_summa': None},
    'ОСП': {'ID': 42, 'cord_summa': None},
    'Москва': {'ID': 61, 'cord_summa': None},
}

for book in dist_book:
    sheet_ranges = wb[book]
    summa_prom = []
    summa = []
    dict_excel = {}
    coord = data()
    coord3 = "A" + coord
    coord4 = "DR" + coord
    printers(dist_book[book])

    for cellObj in sheet_ranges[coord3:coord4]:
        for cell in cellObj:
            coordinate = cell.coordinate
            row = cell.row  # текущий ряд ячейки
            row_2 = row_1(row)  # предыдущий ряд ячейки
            coord_column1 = cell.column  # текущий столбец ячейки
            value_5 = sheet_ranges[cell.column_letter + str(5)].value  # Чтение ячейки столбца типа (счетчик и т.п.)
            if cell.value is None:
                if value_5 == "Счетчик":
                    coord6 = coord_column1 + int(row_2)
                    value_6 = sheet_ranges[cell.column_letter + str(coord6)].value
                    sheet_ranges[coordinate] = value_6  # Запись в ячейку
                elif value_5 == "Прирост":
                    summa = raschet(cell, row_2)
                elif value_5 == "Сумма":
                    for element in summa_prom:
                        summa.append(element)
                    summa_string = '+'.join(summa)
                    summa_string = "=" + summa_string
                    sheet_ranges[coordinate] = summa_string
                    dist_book[book]['cord_summa'] = '=' + sheet_ranges.title + '!' + coordinate
                elif value_5 == "Разница Сумм":
                    raschet(cell, row_2)

    # Удаляем запись none, что бы не баговал
    del dict_excel[None]

    for dict_keys in dict_excel.keys():
        """
            Выставляет Резерв в ячейках.
            Которых нету в zabbix.
        """
        if dict_keys not in serial_zabbix:  # Этот серийник не определяется в Zabbix
            coord7 = dict_excel[dict_keys] + "3"
            coord8 = dict_excel[dict_keys] + "4"
            sheet_ranges[coord7] = "Резерв!"
            sheet_ranges[coord8] = "Резерв!"

            # Копируем занчение из предыдущей ячейки, если нету данный в zabbix
            prev_cord = int(coord) - 1
            last_coord_pages = dict_excel[dict_keys] + coord
            prev_coord_pages = dict_excel[dict_keys] + str(prev_cord)
            sheet_ranges[last_coord_pages].value = sheet_ranges[prev_coord_pages].value

# Собираем лист 'AK'
sheet_ranges = wb['AK']
coord = data()
coord3 = "A" + coord
coord4 = "K" + coord
summa = []
summa_prom = []

for book in dist_book:
    all_value_result(dist_book[book]['cord_summa'])

for cellObj in sheet_ranges[coord3:coord4]:
    for cell in cellObj:
        coordinate = cell.coordinate
        row = cell.row  # текущий ряд ячейки
        row_2 = row_1(row)  # предыдущий ряд ячейки
        coord_column1 = cell.column  # текущий столбец ячейки
        value_2 = sheet_ranges[cell.column_letter + str(2)].value  # Чтение ячейки столбца типа (счетчик и т.п.)
        if cell.value is None:
            if value_2 == "Сумма":
                for element in summa_prom:
                    summa.append(element)
                summa_string = '+'.join(summa)
                summa_string = "=" + summa_string
                sheet_ranges[coordinate] = summa_string
            elif value_2 == "Разница Сумм":
                raschet(cell, row_2)

try:
    wb.save(path)
    log("Запись завершена\n\n")
except PermissionError:
    log("Не возможно сохранить файл excel, у кого-то он открыт! Сохранение в промежуточный файл!\n\n")
    raise

# временно отключено, до запуска в автоматизацию
# try:
#     shutil.copy(path, path_copy)  # !!!!!!!!
# except PermissionError:
#     log("ПЗДЦ!!!! Файл не скопировался, видимо потому что он открыт у кого-то! Копирование в общую папку!\n\n")
#     raise
